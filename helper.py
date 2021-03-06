import openpyxl
from binance import Client
from binance.enums import *
from decimal import *

import config
import ccxt

exchange = ccxt.binance();


def get_historic_data_by_symbol_days(client,sym,days):
    return client.get_historical_klines(sym, Client.KLINE_INTERVAL_1DAY, f'{days} day ago UTC')

def get_historic_data_by_start_and_end_date(client,sym,interval,start_date,end_date):
    return client.get_historical_klines(sym, interval, start_date, end_date)


# API for doing money management
def get_quantity_of_stock_for_buy(buy_price_usdt):
    total_money_for_trade = config.TOTAL_USDT *(1 - config.PARK_RATIO/100)
    total_coin_of_interest = config.NO_OF_COINS_OF_INTEREST
    money_available_per_trade = total_money_for_trade/total_coin_of_interest
    #quantity = round(money_available_per_trade/buy_price_usdt, 5)
    print(money_available_per_trade/buy_price_usdt)
    quantity = format_num_filter(money_available_per_trade/buy_price_usdt)
    print(f' Quantity for this trade is {quantity}')
    return quantity

def format_num_filter(number):
    if 5>number >= 1 :
        return round(number,1)
    elif 1 > number and number>=0.1 :
        return  round(number,2)
    elif 0.1 >number and number>=0.01:
        return round(number,3)
    elif 0.01 > number and number>= 0.001 :
        return round(number,4)
    elif 0.001 > number and number>= 0.0001 :
        return round(number,5)
    elif 0.0001 > number and number>= 0.00001 :
        return round(number,6)
    elif 0.00001 > number and number>= 0.000001 :
        return round(number,7)
    elif 0.000001 > number and number>= 0.0000001 :
        return round(number,8)
    else:
        return round(number)

# API to place GTT buy orders
def buy(client,sym,stopprice):
    #stop_price = round(float(stopprice),2)
    exchange = get_exchange()
    print(exchange.fetch_balance())
    #stop_price = format_num_filter(float(stopprice))
    coin_symbol = sym.split('USDT')[0]+"/USDT"
    stop_price = exchange.price_to_precision(coin_symbol,stopprice)
    if check_if_order_does_not_exists(client,sym) is False :
        print(f'Order exists for {sym} going to cancel the previous order')
        cancel(client, sym)
    #TODO Add a wait of 2 secs
    #quantity_of_stocks_to_buy = format_num_filter(get_quantity_of_stock_for_buy(stop_price *1.005))
    quantity_of_stocks_to_buy = exchange.amount_to_precision(coin_symbol,get_quantity_of_stock_for_buy(float(stop_price) *1.005))
    #price = round((stop_price *1.01),1) #Placing order 1 % above the GTT price
    #price = format_num_filter(stop_price + (stop_price *0.01))  #Placing order 1 % above the GTT price
    #price = stop_price +  format_num_filter((stop_price *0.01))
    price = float(stop_price) +  float(exchange.price_to_precision(coin_symbol,(float(stop_price) *0.01)))
    #print(Decimal(price))#Placing order 1 % above the GTT price
    order = client.create_order(symbol=sym,
                                side=SIDE_BUY,
                                type = ORDER_TYPE_STOP_LOSS_LIMIT,
                                timeInForce= TIME_IN_FORCE_GTC,
                                quantity=exchange.amount_to_precision(coin_symbol,quantity_of_stocks_to_buy),
                                price= exchange.price_to_precision(coin_symbol,price),
                                stopPrice =  stop_price
                                )

    client.create_order


#API to cancel the existing order
def cancel(client,sym):
    order = client.get_open_orders(symbol=sym)
    order_id = order[0].get('orderId')
    #print(order)
    order = client.cancel_order(symbol=sym,orderId = order_id)


#API to check if the order exists
def check_if_order_does_not_exists(client,sym):
    order = client.get_open_orders(symbol=sym)
    if len(order) > 0 :
        print(f' Pending order already exists for {sym}')
        return False
    else :
        print(f' Pending order doesnot exists for {sym}')
        return True

#API to initialize the exchange
exchange = None
def get_exchange():
    # defining the exchange from where we want the data , we are connecting to binance
    # Note : We are not in binanceus
    global exchange
    if exchange is None:
        exchange = ccxt.binance({
            'apiKey': config.API_KEY,
            'secret': config.API_SECRET
        })
    # print(f'Balance {exchange.fetch_balance()}')
    return exchange

# Check in_position for symbol
def check_in_position(coin_symbol, exchange):
    coin = str(coin_symbol).split('USDT')
    balances = exchange.fetch_balance()
    assets = balances.get("info").get("balances")
    #last_row_index = len(df.index) - 1
    for asset in assets:
        if asset.get('asset') == coin[0]:
            free = float(asset['free'])
            locked = float(asset['locked'])
            print(f'Free Balacne is {free} and locked is {locked}')
            # Fetching the latest price of coin_symbol(ETH/USDT)
            price = float(exchange.fetchTicker(coin_symbol).get('last'))
            current_position_size = price * (free + locked)
            if current_position_size > 10:
                print(
                    f'You are already in position for {coin_symbol} , free balance {free} and estimated position size {current_position_size}')
                return True
    return False

#Symbols We are tracking
sym_all_from_main=set()
def fetch_all_symbols_from_main():
    wb = openpyxl.load_workbook('sst_1.xlsx')
    sh1 = wb['Main']
    row = sh1.max_row
    for i in range(1, row):
        sym_all_from_main.add(sh1.cell(i + 1, 1).value)
    print(sym_all_from_main)

#List is in USDT pair . eg: BTCUSDT
sym_where_we_have_position=set()
def fetch_all_sym_where_we_have_positions(exchange):
    balances = exchange.fetch_balance()
    assets = balances.get("info").get("balances")
    for asset in assets:
        free = float(asset['free'])
        locked = float(asset['locked'])
        print(f'Free Balacne is {free} and locked is {locked}')
        # Fetching the latest price of coin_symbol(ETH/USDT)
        coin_symbol=str(asset['asset'])+'USDT'
        if coin_symbol in sym_all_from_main:
            price = float(exchange.fetchTicker(coin_symbol).get('last'))
            current_position_size = price * (free + locked)
            if current_position_size > 10:
                sym_where_we_have_position.add(coin_symbol)
                print(
                    f'You are already in position for {coin_symbol} , '
                    f'free balance {free} and estimated position size {current_position_size}')

    print(sym_where_we_have_position)


def check_in_position_v2(coin_symbol):
    if coin_symbol in sym_where_we_have_position:
        return True
    return False
